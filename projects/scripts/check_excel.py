#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel文件的数据结构
"""
import openpyxl
import sys

def check_excel(excel_path):
    """检查Excel文件的数据结构"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    print(f"\n检查Excel: {excel_path}\n")
    print("=" * 80)
    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}\n")

    # 显示第1行
    print("第1行（表头）:")
    for col_idx, cell in enumerate(ws[1], start=1):
        print(f"  列{col_idx}({chr(64+col_idx)}): {cell.value}")

    print("\n从第10行开始的数据（前20行）:")
    for row_idx in range(10, min(30, ws.max_row + 1)):
        a_value = ws[f'A{row_idx}'].value
        print(f"  行{row_idx}: A列={a_value}")

        if a_value and '未税合计' in str(a_value):
            print(f"  >>> 在行{row_idx}找到'未税合计'，停止读取 <<<")
            break

    print("\n" + "=" * 80)

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("用法: python3 check_excel.py <xlsx文件路径>")
        sys.exit(1)

    check_excel(sys.argv[1])
