#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查Excel第10行开始的详细数据
"""
import openpyxl
import sys

def check_excel_detail(excel_path):
    """检查Excel第10行开始的详细数据"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    print(f"\n检查Excel: {excel_path}\n")
    print("=" * 80)
    print(f"从第10行开始的详细数据:\n")

    # 列映射
    columns = {
        'A': 'BH',
        'C': 'CPMC',
        'D': 'CPMS',
        'E': 'CPSL',
        'F': 'CPDW',
        'G': 'HSDJ',
        'I': 'SL',
        'J': 'WSJE',
        'L': 'HSJE'
    }

    # 显示表头
    print(f"{'行号':<5}", end="")
    for col_name, col_desc in columns.items():
        print(f"{col_name}({col_desc}):<30", end="")
    print()

    # 读取数据
    for row_idx in range(10, ws.max_row + 1):
        a_value = ws[f'A{row_idx}'].value

        if a_value and '未税合计' in str(a_value):
            print(f"{'-'*5}", end="")
            for col_name in columns:
                print(f"{'-'*30}", end="")
            print(f"\n  >>> 在行{row_idx}找到'未税合计'，停止读取 <<<\n")
            break

        print(f"{row_idx:<5}", end="")
        for col_name in columns:
            cell_value = ws[f'{col_name}{row_idx}'].value
            display_value = str(cell_value)[:28] if cell_value else ''
            print(f"{display_value:<30}", end="")
        print()

    print("\n" + "=" * 80)

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("用法: python3 check_excel_detail.py <xlsx文件路径>")
        sys.exit(1)

    check_excel_detail(sys.argv[1])
